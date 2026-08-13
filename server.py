from flask import Flask, request, jsonify, send_from_directory, Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime, timezone
from threading import Lock
import os, json, hashlib, base64
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PublicKey
from cryptography.exceptions import InvalidSignature

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv('DATA_DIR', BASE_DIR / 'data'))
EXCEL_PATH = DATA_DIR / 'inscripciones_congreso.xlsx'
LOCK = Lock()
INTEGRITY_PATH = BASE_DIR / 'integrity.manifest.json'
PUBLIC_KEY_B64 = 'izpzVlUDfMe/ud6hkeVn4ywtRz2dqD4oiDSAbxsdl98='


def verify_integrity():
    try:
        manifest = json.loads(INTEGRITY_PATH.read_text(encoding='utf-8'))
        signature = base64.b64decode(manifest.pop('signature_b64'))
        canonical = json.dumps(manifest, sort_keys=True, separators=(',', ':')).encode()
        Ed25519PublicKey.from_public_bytes(base64.b64decode(PUBLIC_KEY_B64)).verify(signature, canonical)
        for rel, expected in manifest.get('files', {}).items():
            path = BASE_DIR / rel
            if not path.is_file():
                return False, f'Falta archivo protegido: {rel}', manifest
            if hashlib.sha256(path.read_bytes()).hexdigest() != expected:
                return False, f'Integridad alterada: {rel}', manifest
        return True, 'Firma e integridad verificadas', manifest
    except (OSError, ValueError, KeyError, InvalidSignature, json.JSONDecodeError) as exc:
        return False, f'Firma inválida: {exc}', {}


def parse_expiration(manifest):
    value = manifest.get('expires_at')
    if not value:
        return None
    dt = datetime.fromisoformat(value)
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


INTEGRITY_OK, INTEGRITY_MESSAGE, SIGNED_MANIFEST = verify_integrity()
EXPIRATION = parse_expiration(SIGNED_MANIFEST) if INTEGRITY_OK else None


def site_expired():
    return bool(INTEGRITY_OK and EXPIRATION and datetime.now(timezone.utc) >= EXPIRATION.astimezone(timezone.utc))


def expired_page():
    expiry_text = EXPIRATION.strftime('%d/%m/%Y %H:%M') if EXPIRATION else ''
    html = """<!doctype html><html lang='es'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>Sitio inactivo</title><style>html,body{height:100%;margin:0}body{font-family:Arial,sans-serif;background:#3C235F;color:#fff;display:grid;place-items:center;padding:24px;box-sizing:border-box}main{max-width:720px;text-align:center}h1{font-size:clamp(2rem,6vw,4.6rem);margin:0 0 18px}p{font-size:1.05rem;line-height:1.6;opacity:.88}small{opacity:.55}</style></head><body><main><h1>Sitio inactivo</h1><p>La vigencia autorizada de esta publicación ha finalizado.</p><small>Vigencia hasta: %s · Ecuador</small></main></body></html>""" % expiry_text
    return Response(html, status=410, mimetype='text/html')


app = Flask(__name__, static_folder='.', static_url_path='')
GENERAL_HEADERS = ['Fecha registro','Nombres','Apellidos','Correo','Teléfono','País','Ciudad','Empresa / Institución','Cargo / Rol','Perfil','Acepta datos','Acepta imagen']
STARTUP_HEADERS = ['Fecha registro','Startup','Sitio web','Fundador / Representante','Correo','Teléfono','País','Sector','Etapa','Año creación','Tamaño equipo','Busca inversión','Capital buscado USD','Descripción','Tracción / métricas','Objetivo en el Congreso','Pitch deck','Acepta datos','Acepta imagen']


def ensure_workbook():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists(): return
    wb = Workbook(); ws = wb.active; ws.title = 'Asistentes'; ws.append(GENERAL_HEADERS)
    ws2 = wb.create_sheet('Startups'); ws2.append(STARTUP_HEADERS)
    for sheet in (ws, ws2):
        for cell in sheet[1]:
            cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='EDE9FE'); cell.alignment = Alignment(horizontal='center')
        sheet.freeze_panes = 'A2'; sheet.auto_filter.ref = sheet.dimensions
    wb.save(EXCEL_PATH)


def clean(v): return '' if v is None else str(v).strip()


def append_row(sheet_name, headers, row):
    with LOCK:
        ensure_workbook(); wb = load_workbook(EXCEL_PATH); ws = wb[sheet_name]; ws.append(row)
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{ws.max_row}"
        for i, header in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(len(header) + 3, 14), 40)
        wb.save(EXCEL_PATH)


@app.before_request
def enforce_signature_and_expiry():
    if request.path == '/health': return None
    if not INTEGRITY_OK: return jsonify({'status':'blocked','reason':INTEGRITY_MESSAGE}), 503
    if site_expired():
        if request.path.startswith('/api/'):
            return jsonify({'ok':False,'status':'expired','message':'La vigencia del sitio ha finalizado.'}), 410
        return expired_page()
    return None


@app.get('/')
def index(): return send_from_directory(BASE_DIR, 'index.html')


@app.post('/api/inscripciones/general')
def register_general():
    data = request.get_json(silent=True) or {}
    required = ['nombres','apellidos','email','telefono','pais','ciudad','perfil']
    missing = [f for f in required if not clean(data.get(f))]
    if missing: return jsonify({'ok':False,'message':'Faltan campos obligatorios.','fields':missing}), 400
    if not data.get('acepta_datos') or not data.get('acepta_imagen'):
        return jsonify({'ok':False,'message':'Debes aceptar los consentimientos obligatorios.'}), 400
    row=[datetime.now().strftime('%Y-%m-%d %H:%M:%S'),clean(data.get('nombres')),clean(data.get('apellidos')),clean(data.get('email')),clean(data.get('telefono')),clean(data.get('pais')),clean(data.get('ciudad')),clean(data.get('organizacion')),clean(data.get('cargo')),clean(data.get('perfil')),'Sí','Sí']
    append_row('Asistentes', GENERAL_HEADERS, row)
    return jsonify({'ok':True,'message':'Inscripción registrada correctamente.'})


@app.post('/api/inscripciones/startup')
def register_startup():
    data = request.get_json(silent=True) or {}
    required = ['startup','fundador','email','telefono','pais','sector','etapa','busca_inversion','descripcion','objetivo']
    missing = [f for f in required if not clean(data.get(f))]
    if missing: return jsonify({'ok':False,'message':'Faltan campos obligatorios.','fields':missing}), 400
    if not data.get('acepta_datos') or not data.get('acepta_imagen'):
        return jsonify({'ok':False,'message':'Debes aceptar los consentimientos obligatorios.'}), 400
    row=[datetime.now().strftime('%Y-%m-%d %H:%M:%S'),clean(data.get('startup')),clean(data.get('web')),clean(data.get('fundador')),clean(data.get('email')),clean(data.get('telefono')),clean(data.get('pais')),clean(data.get('sector')),clean(data.get('etapa')),clean(data.get('anio')),clean(data.get('equipo')),clean(data.get('busca_inversion')),clean(data.get('capital')),clean(data.get('descripcion')),clean(data.get('traccion')),clean(data.get('objetivo')),clean(data.get('pitch')),'Sí','Sí']
    append_row('Startups', STARTUP_HEADERS, row)
    return jsonify({'ok':True,'message':'Postulación registrada correctamente.'})


@app.get('/health')
def health():
    if not INTEGRITY_OK:
        return jsonify({'status':'tampered','integrity':False,'detail':INTEGRITY_MESSAGE}), 503
    expired = site_expired()
    return jsonify({'status':'expired' if expired else 'ok','integrity':True,'expired':expired,'expires_at':SIGNED_MANIFEST.get('expires_at'),'detail':'Firma válida; vigencia finalizada.' if expired else 'Firma e integridad verificadas.'}), (410 if expired else 200)


if __name__ == '__main__':
    ensure_workbook(); app.run(host='0.0.0.0', port=8080)
